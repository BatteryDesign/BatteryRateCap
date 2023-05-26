"""
This module is used to convert battery cycling data
and battery charge/discharge data into capacity-rate data.
"""
import os
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.mixture import GaussianMixture


def potential_rate_paper_set(input_file, sheet_name,
                             output_file, paper_num, set_num):
    """
    This function converts potential vs. capacity data to capacity vs c-rate
    data. The Highest x-value (capacity [mAh]) from the charge/discharge graph.
    The input excel file has the format that each excel file has many sheets
    and each sheet, separated by charge/discharge C-rate, contains many papers.
    Each paper can have more than 1 set of capacity versus voltage data, each
    occupying 1 column (i.e. each set takes up two columns). The first row of
    a sheet specifies the paper code number, e.g. 'paper # 1', and the second
    row of a sheet spcifies the battery's set number, e.g. 'set # 1', and the
    third row of a sheet contains the quantity name and units, e.g. 'Capacity
    (mAh/g)' and 'Voltage (V)'.
    PARAMETERS
    ----------
    1) Excel file (file path) - string
    2) sheetnames - list
    3) Paper # - string
    4) Set # - integer: max number of sets
    RETURNS
    -------
    Capacity vs c-rate dataframe
    """
    # Determining which set # and the number of set lists from the excel file
    set_list = ['set #' + str(i) for i in range(1, set_num + 1)]
    # Assert that the input 'set_num' is an integer
    assert isinstance(set_num, int) is True, 'set_num must be an integer'
    # Assert that the input 'paper_num' is a string and has the correct format
    assert isinstance(paper_num, str) is True, 'paper_num must be a string'
    assert 'Paper # ' in paper_num, 'paper_num not in correct format'
    # Assert input sheet names are in a list
    if isinstance(sheet_name, list) is not True:
        raise TypeError('sheet_name must be a list of strings')
    # Assert sheetname contains 'C_'
    for _, name in enumerate(sheet_name):
        assert 'C_' in name, 'input sheet names are not in the correct format'
    # Dataframing the interested potential vs capacity excel sheet
    df_input = pd.read_excel(input_file, sheet_name, header=[0, 1, 2])
    # Merging multiple spreadsheets
    df_sheets = []
    for _, name in enumerate(sheet_name):
        df_sheets.append(df_input[name])
    df_merged = pd.concat(df_sheets, axis=1)
    # Get C-rate numerical values
    rates = []
    for name in sheet_name:
        rate = name.split("C_")[0]
        rates.append(rate)
    c_rate = pd.DataFrame({"C rate": rates})
    # Selecting maximum capacity values for each dataset and
    # concatnate with corresponding c-rates or current density
    caplist = []
    for i in set_list:
        set_i = (df_merged[paper_num, i])
        set_i_max = set_i["Capacity (mAh/g)"].max(axis=0)
        # if there is only one capcacity result
        if isinstance(set_i_max, float):
            set_i_max = pd.array([set_i_max])
        else:
            set_i_max = set_i_max.array
        caplist.append(c_rate)
        caplist.append(pd.DataFrame({"Capacity (mAh/g)": set_i_max}))
    df_cap_rate = pd.concat(caplist, axis=1)
    # Test that the output is a dataframe
    assert isinstance(df_cap_rate, pd.DataFrame) is True, '\
    The output must be a dataframe'
    # Exporting the converted dataframe to an excel file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in dataframe_to_rows(df_cap_rate, index=True, header=True):
        worksheet.append(row)
    workbook.save(output_file)
#     df_cap_rate.to_excel(output_file,sheet_name=paper_num,
#                          index=False, header=True)
    # Test that the sheet name is a string
    assert isinstance(paper_num, str) is True, 'sheetname must be a string'
    print('saved succesfully to' + output_file)
    return df_cap_rate


def potential_rate_all(input_file, output_file):
    """
    This function converts and dataframes all voltage potential data and
    separates them by paper and set numbers such that the users can see
    what options they have. The input excel file has the format that each
    excel file has many sheets and each sheet, separated by charge/discharge
    C-rate, contains many papers. Each paper can have more than 1 set of
    capacity versus voltage data, each occupying 1 column (i.e. each set
    takes up two columns). The first row of a sheet specifies the paper code
    number, e.g. 'paper # 1', and the second row of a sheet spcifies the
    battery's set number, e.g. 'set # 1', and the third row of a sheet
    contains the quantity name and units, e.g. 'Capacity (mAh/g)' and
    'Voltage (V)'.
    PARAMETERS
    ----------
    1) Excel file (file path) - string
    2) Output Excel file name - string
    RETURNS
    -------
    A voltage-capaity dataframe by paper and set number
    """
    # Read all sheets in the input file into a dictionary
    dict_excel = pd.read_excel(input_file, sheet_name=None, header=[0, 1, 2])
    sheetnames = list(dict_excel.keys())
#     dict_excel[sheetnames[0]]
    df_cap_rate = pd.DataFrame()
    for _, sheetname in enumerate(sheetnames):
        df_input = dict_excel[sheetname]
        rate = sheetname.split("C_")[0]
        for headers, columnval in df_input.items():
            paper_num, set_num, quan = headers
            assert 'Paper' in paper_num, 'Wrong paper number header'
            assert 'set' in set_num, 'Wrong set number format'
            # Takes only the capacity data
            if 'capacity' in quan or 'Capacity' in quan:
                max_cap = np.nanmax(columnval.values)
                # Must create the dataframe first
                # before adding multiple headers
                newdf = pd.DataFrame([[rate, max_cap]])
                # Add headers
                newdf_header = [[paper_num, paper_num],
                                [set_num, set_num], ['C-rate', quan]]
                newdf.columns = newdf_header
                df_cap_rate = pd.concat([df_cap_rate, newdf])
    # apply lambda function to each column to drop Nans
    df_cap_rate_all = df_cap_rate.apply(lambda x: pd.Series(x.dropna().values))
    # Export dataframe to Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in dataframe_to_rows(df_cap_rate_all, index=True, header=True):
        worksheet.append(row)
    workbook.save(output_file)
#     df_cap_rate_all.to_excel(output_file,sheet_name='CapacityRate',
#                              index=True, header=True)
    return df_cap_rate_all


def excel_merge(dataframe, xls_file, sheetname):
    """
    This function adds the converted dataframe to an existing excel file
    PARAMETERS
    ----------
    1) converted dataframe - dataframe
    2) Excel file (file path) - string
    3) sheetname - string
    RETURNS
    -------
    A new sheet in the excel file
    """
    # Exporting the converted dataframe to an excel file
    filename, ext = os.path.splitext(xls_file)
    assert ext == '.xls' or ext == '.xlsx', 'Wrong output file type'
    # Test that the sheet name is a string
    assert isinstance(sheetname, str) is True, 'sheetname must be a string'
    dataframe.to_excel(xls_file, sheet_name=sheetname,
                       index=False, header=True)
    print('saved succesfully to' + xls_file)


def capacity_cycle(capacity_cycle_array, num_rate,
                   current_list, current_unit, capacity_unit):
    '''
    This function converts capacity-cycle data to capacity-rate plots.
    For different selections of paper/set, the user will have to adjust
    'n_components' until the groups are properly grouped, as seen in the
    plot below.
    Inputs
    - capacity_cycle_array: nx2 capacity vs cycle # array
    - n: number of C-rates, or 'stairs'
    - current_list: a list of current rates or current densities
    - current_unit: a text string of current unit
    - capacity_unit: a text string od capacity unit
    '''
    assert capacity_cycle_array.shape[1] == 2, 'Input array wrong shape'
    assert len(current_list) == num_rate, 'Input number of current wrong'
    model = GaussianMixture(n_components=num_rate)
    model.fit(capacity_cycle_array)
    # Use the model to make predictions about which group each datapoint
    # belongs to.
    # Predictions stored as an np array with indexes corresponding to points,
    # and values to their assigned class
    prediction = model.predict(capacity_cycle_array)
    # np.array of the unique classes
    clusters = np.unique(prediction)
    # Plot the points now that they are grouped
    plt.figure(figsize=(10, 8))
    for cluster in clusters:
        row_ix = np.where(prediction == cluster)
        plt.scatter(capacity_cycle_array[row_ix, 0],
                    capacity_cycle_array[row_ix, 1],
                    s=300)
    plt.ylabel("Capacity "+capacity_unit, fontsize=22)
    plt.xlabel("Cycle #", fontsize=22)
    # Return the means of each 'stair' and sort
    means = model.means_
    means = means[np.argsort(means[:, 0])]
    means_of_groups = means[:, 1]
    means_of_groups = pd.DataFrame(means_of_groups)
    means_of_groups = means_of_groups.rename(
        columns={0: "Capacity"+capacity_unit})
    # Get list of current rates or current densities
    current_list = pd.DataFrame(np.array(current_list))
    current_header = "Current " + current_unit
    current_list = current_list.rename(columns={0: current_header})
    capacity_vs_current_density_df = pd.concat([current_list,
                                                means_of_groups], axis=1)
    return capacity_vs_current_density_df
